#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
荣誉战场发奖数据
前100名根据排名和分数发奖
超过100名根据分数发奖
前100还有勋章奖励
"""
# import os
# import datetime
# from openpyxl import load_workbook
import time
import configparser
from openpyxl import Workbook
import get_sql

reward = {
    1: 300000,
    2: 250000,
    3: 200000,
    4: 160000,
    5: 120000,
    6: 90000,
    7: 60000,
    8: 30000,
}
reward_title = {
    1: "大元帅",
    2: "元帅",
    3: "大将军",
    4: "将军",
    5: "大都统",
    6: "都统",
    7: "大统领",
    8: "统领",
}

# 获取时间 转成对应的格式
def get_time():
    t1 = time.time()
    time_local = time.localtime(t1)
    t2 = time.strftime("%Y-%m-%d %H:%M:%S", time_local)
    return t2


# 记录log
def write_log(log):
    print(log)
    with open("荣誉战场奖励log.txt", "a", encoding="utf-8") as logfile:

        logfile.write(str(log) + "\n")


# 读取ini
def read_ini():
    write_log("read_ini ... ")
    db_name = []
    try:
        write_log("read ini start...")
        config = configparser.ConfigParser()
        config.read("rongyuzhanchang.ini", encoding="utf-8")

        excel_path = config.get("sql_path", "excel_path")
        script_path = config.get("sql_path", "script_path")
        column_name = config.get("column_name", "column")
        sql = config.get("sql_select", "sql_ini")
        season = config.get("column_name", "season")

        for i in range(1, 12):
            get_dbname = config.get("dbname", str(i)) + season
            write_log("read_ini 编号 {0} ； 区服信息： {1}".format(i,get_dbname))
            db_name.append(get_dbname)

        return (db_name, column_name, sql, excel_path, script_path)
    except:
        write_log("read ini error...")
        print("err")


##can support excel 2007 ,support xlsx.  becouse .xls only support 65535 row


def write_excel_xlsx(dbname, column_name, sql, excel_path, script_path):
    wb = Workbook()
    sheet = wb.active

    result = get_sql.connect_mysql(dbname, sql)

    fileds = column_name.split(",")
    idnum = 0

    dataList1 = []
    dataList_top100 = []
    try:
        write_log("判断排名")
        sheet.append(fileds)
        for row in range(1, len(result) + 1):
            idnum = idnum + 1
            sheet.cell(row=row + 1, column=1, value=idnum)
            userid = result[row - 1][0]
            # print(userid)
            area_id = result[row - 1][3]
            if idnum <= 100:
                dataList_top100.append(userid)
                dataList_top100.append(idnum)

                sheet.cell(row=row + 1, column=12, value=idnum)
                sheet.cell(row=row + 1, column=13, value=userid)
                attr = order_expand_attr(idnum)
                sheet.cell(row=row + 1, column=14, value=attr)

            # print(area_id)
            for col in range(0, len(fileds) - 1):
                sheet.cell(row=row + 1, column=col + 2, value=result[row - 1][col])
                try:
                    # write_log("判断排名")
                    if col == 2:
                        order = order_num(idnum, result[row - 1][col])
                        reward_num = reward.get(order)
                        reward_title_txt = reward_title.get(order)
                        sheet.cell(row=row + 1, column=7, value=order)
                        sheet.cell(row=row + 1, column=8, value=reward_num)
                        sheet.cell(row=row + 1, column=9, value=reward_title_txt)
                        sheet.cell(row=row + 1, column=10, value=area_id)

                        # write_log(sql_num)
                        dataList1.append(userid)
                        dataList1.append(area_id)
                        dataList1.append(reward_num)
                except:
                    write_log("error! 判断排名  error")
    except:
        write_log("error! write excel error")

    wb.save(excel_path + dbname + ".xlsx")
    write_log("保存Excel成功")

    mail_dbname = dbname+"_mail"
    write_top100_sql(script_path, mail_dbname, dataList_top100)
    write_sql(script_path, dbname, dataList1)

# 路径 文件名 数据，保存前100的sql，主要是玩家杯子
def write_top100_sql(script_path, filename, dataList1):
    write_log("write write_top100_sql start...")
    with open(script_path + filename + ".sql", "a", encoding="ansi") as f1:

        f1.write("set names gbk;\n")
        try:
            for strlist in range(0, int(len(dataList1) / 2)):
                userid = dataList1[strlist * 2]
                idnum = dataList1[strlist * 2 + 1]
                expand_attr = order_expand_attr(idnum)
                k = (
                    'insert into mails (tar_user_id,prop,title,content,send_time,due_time,item1,item1_prop,item2,item2_prop,item3,item3_prop,item4,item4_prop,item5,item5_prop) values('
                    + str(userid)
                    +',1,"荣誉战场军衔排名奖励","亲爱的英魂玩家，恭喜您上赛季荣誉战场军衔排名第'
                    +str(idnum)
                    +'名,请查收您的奖励！",2103300000,1624636800,'
                    + str(expand_attr)
                    +',3246,0,0,0,0,0,0,0,0);\n'
                )

                f1.write(k)
        except:
            write_log("写write_top100_sql错误")


def write_sql(script_path, filename, dataList1):
    write_log("write sql start...")
    with open(script_path + filename + ".sql", "a", encoding="ansi") as f1:
        write_log(script_path + filename + ".sql")
        try:
            for strlist in range(0, int(len(dataList1) / 3)):
                userid = dataList1[strlist * 3]
                area_id = dataList1[strlist * 3 + 1]
                reward = dataList1[strlist * 3 + 2]

                fenbiao = int(int(userid) / 1000000)
                k = (
                    "update `pve_role_list"
                    + str(fenbiao)
                    + "` set pve_hornor = pve_hornor + "
                    + str(reward)
                    + " where role_id =  "
                    + str(userid)
                    + " and area_id =  "
                    + str(area_id)
                    + " limit 1;"
                    + "\n"
                )
                f1.write(k)
        except:
            write_log("写sql错误")

        # write_log("close f1")
        # f1.close


def order_num(num, score):
    if score >= 400000:
        if num < 11:
            return 1
        elif num < 101:
            return 2
        else:
            return 3
    elif score >= 350000:
        if num < 101:
            return 2
        else:
            return 3
    elif score >= 310000:
        return 3
    elif score >= 270000:
        return 4
    elif score >= 230000:
        return 5
    elif score >= 200000:
        return 6
    elif score >= 170000:
        return 7
    else:
        return 8

'''
def order_expand_attr(num):
    if num == 1:
        return 131072
    elif num == 2:
        return 262144
    elif num == 3:
        return 524288
    elif num < 11:
        return 1048576
    elif num < 51:
        return 2097152
    elif num < 101:
        return 4194304
    else:
        write_log("判断 expand_attr 错误")
'''
# 判断给的杯子
def order_expand_attr(num):
    if num == 1:
        return 1401
    elif num == 2:
        return 1402
    elif num == 3:
        return 1403
    elif num < 11:
        return 1404
    elif num < 51:
        return 1405
    elif num < 101:
        return 1406
    else:
        write_log("判断 expand_attr 错误")


'''
1401		战场军衔排行第一
1402		战场军衔排行第二
1403		战场军衔排行第三
1404		战场军衔排行十强
1405		战场军衔排行五十强
1406		战场军衔排行百强

'''


if __name__ == "__main__":
    write_log("--------------{0} 开始执行--------------".format(get_time()))
    get_ini = read_ini()
    for db_name in get_ini[0]:

        write_excel_xlsx(db_name, get_ini[1], get_ini[2], get_ini[3], get_ini[4])

    write_log("--------------{0} 执行完成--------------".format(get_time()))

