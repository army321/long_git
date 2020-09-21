#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
pvp排名奖励导出
导出各服的pvp排名前100到同一个文件中
导出各服的排序到单个文件中
"""
import os
import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import datetime
import configparser

# 记录log
def write_log(log):
    logfile = open("D:/python/pvplog.txt", "a", encoding="utf-8")

    t1 = time.time()
    time_local = time.localtime(t1)
    t2 = time.strftime("%Y-%m-%d %H:%M:%S", time_local)

    logfile.write(t2 + " : " + log + "\n")
    logfile.close


def connect_mysql(dbname, chuan_sql):
    config = {
        "host": "127.0.0.1",
        "port": 3306,
        "user": "root",
        "passwd": "123456",
        "db": dbname,
    }

    con = pymysql.connect(**config)
    cur = con.cursor()
    sql = chuan_sql

    try:
        cur.execute(sql)
        result = cur.fetchall()
        return result
    except:
        print("error!  select mysql error")

    con.close()


def read_ini():
    print("read_ini")

    try:
        write_log("read ini start...")
        config = configparser.ConfigParser()
        config.read("D:/python/pvp_reward.ini", encoding="utf-8")

        excel_path = config.get("sql_path", "excel_path")
        script_path = config.get("sql_path", "script_path")
        column_name = config.get("column_name", "column")
        sqlfz = config.get("sql_select", "sql_fz_top100")
        sqljz = config.get("sql_select", "sql_jz_top100")
        top_file = config.get("sql_path", "top_file_path")

        for i in range(1, 12):
            print("read_ini111")
            mode = "fz"
            get_dbname = config.get("dbname", str(i))
            print(
                get_dbname
                + ","
                + get_dbname
                + ","
                + column_name
                + ","
                + sqlfz
                + ","
                + excel_path
                + ","
                + script_path
                + ","
                + mode
                + ","
                + top_file
            )
            write_excel_xlsx(
                get_dbname,
                get_dbname,
                column_name,
                sqlfz,
                excel_path,
                script_path,
                mode,
                top_file,
            )

        for i in range(1, 12):
            print("read_ini111")
            mode = "jz"
            get_dbname = config.get("dbname", str(i))
            print(get_dbname)
            write_excel_xlsx(
                get_dbname,
                get_dbname,
                column_name,
                sqljz,
                excel_path,
                script_path,
                mode,
                top_file,
            )

    except:
        write_log("read ini error...")
        print("err")


def write_excel_xlsx(
    dbname, file_name, column_name, sql, excel_path, script_path, mode, top_file
):
    print("START EXCEL")
    wb = load_workbook(top_file)
    # 这边要先创建top100.xlsx 的空文件

    sheet = wb.create_sheet(file_name + mode)
    # sheet = wb.active

    print("START EXCEL111")
    result = connect_mysql(dbname, sql)
    print("START result")
    fileds = column_name.split(",")
    idnum = 0

    # dataList1 = []
    # dataList_top100 = []
    try:
        write_log("判断排名")
        sheet.append(fileds)
        for row in range(1, len(result) + 1):
            idnum = idnum + 1
            sheet.cell(row=row + 1, column=1, value=idnum)

            for col in range(0, len(fileds) - 1):
                sheet.cell(row=row + 1, column=col + 2, value=result[row - 1][col])

    except:
        print("error! write excel error")

    wb.save(top_file)
    print("baocun cheng gong")


if __name__ == "__main__":
    print("start ......")

    read_ini()

    print("zhixing wan cheng")

