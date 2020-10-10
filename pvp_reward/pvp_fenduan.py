#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
根据分段发放奖励
根据分段发放杯子
根据当前分数发放前100杯子
所有sql整合成最终的

"""
import os
import time
import datetime
import configparser

from openpyxl import Workbook
from openpyxl import load_workbook


import get_sql


reward = {1: 2800, 2: 2500, 3: 2200}
reward_beizi = {1: 2400, 2: 2600, 3: 2800}


mail_title = "insert into mails (tar_user_id,prop,title,content,send_time,due_time,item1,item1_prop,item2,item2_prop,item3,item3_prop,item4,item4_prop,item5,item5_prop) values"
sql_beizi = "update role_list set expand_attr= expand_attr | "


def read_ini():
    print("read_ini")
    get_dbname_all = []
    global season
    global send_time
    global rw_2800
    global rw_2500
    global rw_2200
    global season_medal
    global QUFU_NUM
    try:
        write_log("read ini start...")
        config = configparser.ConfigParser()
        config.read("pvp_fenduan.ini", encoding="utf-8")

        excel_path = config.get("sql_path", "excel_path")
        script_path = config.get("sql_path", "script_path")
        column_name = config.get("column_name", "column")
        sql_fz = config.get("sql_select", "sql_ini_fz")
        sql_jz = config.get("sql_select", "sql_ini_jz")

        sql_top_fz = config.get("sql_select_top100", "sql_fz_top100")
        sql_top_jz = config.get("sql_select_top100", "sql_jz_top100")
        column_top_name = config.get("column_name", "column_top")

        season = config.get("season_reward", "season")
        send_time = config.get("season_reward", "send_time")

        rw_2800 = config.get("season_reward", "rw_2800")
        rw_2500 = config.get("season_reward", "rw_2500")
        rw_2200 = config.get("season_reward", "rw_2200")

        season_medal = config.get("season_reward", "season_medal")
        QUFU_NUM = config.get("season_reward", "QUFU_NUM")

        for i in range(1, int(QUFU_NUM)):
            get_dbname = config.get("dbname", str(i))
            print(get_dbname)
            get_dbname_all.append(get_dbname)

        return (
            get_dbname_all,
            column_name,
            excel_path,
            script_path,
            sql_fz,
            sql_jz,
            sql_top_fz,
            sql_top_jz,
            column_top_name,
        )
    except:
        write_log("read ini error...")


def get_reward_sql(userid, score_level, gamemode, reward_level):

    if gamemode == 1101:
        mode = "纷争圣坛"
    elif gamemode == 1102:
        mode = "决战之谷"

    return '({0},1,"客服中心","亲爱的英魂玩家,恭喜您在{1}赛季{2}达到{3}分,现已将您的奖励发入,请您查收！",{4}{5}{6})'.format(
        userid, season, mode, score_level, send_time, reward_level, season_medal
    )


def order_reward(tar_user_id, score, gamemode):

    if score >= 2800:
        return get_reward_sql(tar_user_id, 2800, gamemode, rw_2800)
    elif score >= 2500:
        return get_reward_sql(tar_user_id, 2500, gamemode, rw_2500)
    elif score >= 2200:
        return get_reward_sql(tar_user_id, 2200, gamemode, rw_2200)
    else:
        write_log("玩家：{0}，模式：{2}，分数：{1} ".format(tar_user_id, score, gamemode))
        pass
        # return None


def get_expand_attr(tar_user_id, expand_attr_id):
    return "{0} {1} where id ={2} limit 1;".format(
        sql_beizi, expand_attr_id, tar_user_id
    )


def get_expand_attr_id(score, gamemode):
    if gamemode == 1102:
        if score >= 2800:
            return 6144
        elif score >= 2600:
            return 4096
        elif score >= 2400:
            return 2048

    if gamemode == 1101:
        if score >= 2800:
            return 192
        elif score >= 2600:
            return 128
        elif score >= 2400:
            return 64


def get_top_expand_attr_id(rank, gamemode):
    if gamemode == 1102:
        if rank == 1:
            return 256
        elif rank == 2:
            return 512
        elif rank == 3:
            return 768
        elif rank <= 10:
            return 1024
        elif rank <= 50:
            return 1280
        elif rank <= 100:
            return 1536

    elif gamemode == 1101:
        if rank == 1:
            return 8
        elif rank == 2:
            return 16
        elif rank == 3:
            return 24
        elif rank <= 10:
            return 32
        elif rank <= 50:
            return 40
        elif rank <= 100:
            return 48


def order_expand_attr(tar_user_id, score, gamemode):
    if score >= 2400:
        return get_expand_attr(tar_user_id, get_expand_attr_id(score, gamemode))


def top_100_expand_attr(tar_user_id, rank, gamemode):
    return get_expand_attr(tar_user_id, get_top_expand_attr_id(rank, gamemode))


# 获取时间 转成对应的格式
def get_time():
    t1 = time.time()
    time_local = time.localtime(t1)
    t2 = time.strftime("%Y-%m-%d %H:%M:%S", time_local)
    return t2


# 记录log
def write_log(log):
    print(log)
    with open("赛季分段奖励log.txt", "a", encoding="utf-8") as logfile:

        logfile.write(str(log) + "\n")


##can support excel 2007 ,support xlsx.  becouse .xls only support 65535 row

# 写分段的奖励数据
def write_excel_xlsx(
    dbname, file_name, column_name, sql, excel_path, script_path, gamemode
):

    wb = Workbook()
    sheet = wb.active
    result = get_sql.connect_mysql(dbname, sql)
    fileds = column_name.split(",")
    reward_list = []
    expand_attr_list = []

    try:
        write_log("分段")
        sheet.append(fileds)
        for row in range(1, len(result) + 1):
            # print(area_id)
            for col in range(0, len(fileds)):
                sheet.cell(row=row + 1, column=col + 1, value=result[row - 1][col])

                try:
                    # write_log("判断排名")
                    if col == 3:
                        tar_user_id = result[row - 1][0]
                        gamemode = result[row - 1][1]
                        score = result[row - 1][3]
                        if score >= 2200:
                            reward_sql = order_reward(tar_user_id, score, gamemode)

                        expand_attr_sql = order_expand_attr(
                            tar_user_id, score, gamemode
                        )

                        sheet.cell(row=row + 1, column=9, value=reward_sql)
                        sheet.cell(row=row + 1, column=7, value=expand_attr_sql)

                        # write_log(sql_num)
                        reward_list.append(reward_sql)
                        if expand_attr_sql != "":
                            expand_attr_list.append(expand_attr_sql)

                except:
                    write_log("error! 判断排名  error")
    except:
        write_log("error! write excel error")

    if gamemode == 1101:
        wb.save("{0}纷争_{1}.xlsx".format(excel_path, file_name))
        write_log("{0}/纷争_{1}.xlsx 保存成功".format(excel_path, file_name))
    elif gamemode == 1102:
        wb.save("{0}决战_{1}.xlsx".format(excel_path, file_name))
        write_log("{0}决战_{1}.xlsx 保存成功".format(excel_path, file_name))
    else:
        write_log("保存分段Excel文件异常")

    reward_list_duankou = "3308"
    expand_attr_list_duankou = "3306"

    new_file_path = "{}/分段奖励/".format(script_path)
    # 如果没有文件夹，则创建一个文件夹用于存放生成后的数据
    if not os.path.exists(new_file_path):
        os.makedirs(new_file_path)

    beizi_file_path = "{}/杯子奖励/".format(script_path)
    # 如果没有文件夹，则创建一个文件夹用于存放生成后的数据
    if not os.path.exists(beizi_file_path):
        os.makedirs(beizi_file_path)

    write_mails_sql(new_file_path, file_name, reward_list, reward_list_duankou)
    write_sql(beizi_file_path, file_name, expand_attr_list, expand_attr_list_duankou)


# 写前100的杯子奖励数据 只有杯子
def write_top_excel_xlsx(
    dbname, file_name, column_name, sql, excel_path, script_path, gamemode
):

    wb = Workbook()
    sheet = wb.active
    result = get_sql.connect_mysql(dbname, sql)
    fileds = column_name.split(",")
    expand_attr_list = []

    try:
        write_log("分段")
        sheet.append(fileds)
        for row in range(1, len(result) + 1):
            # print(area_id)
            for col in range(0, len(fileds)):
                sheet.cell(row=row + 1, column=col + 1, value=result[row - 1][col])

                try:
                    # write_log("判断排名")
                    if col == 4:
                        tar_user_id = result[row - 1][0]
                        gamemode = result[row - 1][1]
                        score = result[row - 1][4]
                        expand_attr_sql = top_100_expand_attr(
                            tar_user_id, row, gamemode
                        )

                        sheet.cell(row=row + 1, column=10, value=expand_attr_sql)

                        if expand_attr_sql != "":
                            expand_attr_list.append(expand_attr_sql)

                except:
                    write_log("error! 判断排名  error")
    except:
        write_log("error! write excel error")

    top_file_path = "{}/前100杯子Excel/".format(script_path)
    # 如果没有文件夹，则创建一个文件夹用于存放生成后的数据
    if not os.path.exists(top_file_path):
        os.makedirs(top_file_path)

    if gamemode == 1101:
        wb.save("{0}纷争TOP100_{1}.xlsx".format(top_file_path, file_name))
        write_log("{0}纷争TOP100_{1}.xlsx 保存成功".format(top_file_path, file_name))
    elif gamemode == 1102:
        wb.save("{0}决战TOP100_{1}.xlsx".format(top_file_path, file_name))
        write_log("{0}决战TOP100_{1}.xlsx 保存成功".format(top_file_path, file_name))
    else:
        write_log("保存分段Excel文件异常")

    beizi_file_path = "{}/杯子奖励/".format(script_path)
    # 如果没有文件夹，则创建一个文件夹用于存放生成后的数据
    if not os.path.exists(beizi_file_path):
        os.makedirs(beizi_file_path)
    expand_attr_list_duankou = "3306"
    write_sql(beizi_file_path, file_name, expand_attr_list, expand_attr_list_duankou)


def write_mails_sql(script_path, filename, dataList1, duankou):
    write_log("write sql start...")

    # f1 = open("{}{}_{}.sql".format(script_path , filename , duankou ), "a", encoding="ansi")
    with open(
        "{}{}_{}.sql".format(script_path, filename, duankou), "a", encoding="ansi"
    ) as f1:
        f1.write("\n" + "set names gbk;" + "\n")
        # f1.write(mail_title)

        write_log("{}{}_{}.sql".format(script_path, filename, duankou))
        # print(filename)
        # 	print(dataList1)
        index_num = 0
        list_len = len(dataList1)
        print(len(dataList1))
        # 这边每1万个分成一个插入段
        # try:
        for strlist in dataList1:
            if index_num % 10000 == 9999:
                index_num += 1
                f1.write("\n" + strlist + ";")
            elif index_num % 10000 == 0:
                f1.write("\n" + mail_title)
                f1.write("\n" + strlist + ",")
                index_num += 1
            else:
                index_num += 1
                if index_num == list_len:
                    f1.write("\n" + strlist + ";")  # 判断最后一行加分号
                else:
                    f1.write("\n" + strlist + ",")

        # except:
        #     write_log("写sql错误")

        write_log("close f1")
        # f1.close


def write_sql(script_path, filename, dataList1, duankou):
    write_log("write sql start...")
    with open(
        script_path + filename + "_" + duankou + ".sql", "a", encoding="ansi"
    ) as f1:
        print(script_path + filename + ".sql")

        # try:
        for strlist in dataList1:
            if strlist != None:
                f1.write("\n" + strlist)
        f1.write("\n \n")
        # except:
        #     write_log("写sql错误")
        write_log("close f1")
        # f1.close


if __name__ == "__main__":
    # read_ini()
    # read_excel()
    get_ini = read_ini()
    # write_log(str(get_ini))

    # 纷争top100杯子
    for db_name in get_ini[0]:
        write_top_excel_xlsx(
            db_name, db_name, get_ini[8], get_ini[6], get_ini[2], get_ini[3], 1101
        )

    # 纷争
    for db_name in get_ini[0]:
        write_excel_xlsx(
            db_name, db_name, get_ini[1], get_ini[4], get_ini[2], get_ini[3], 1101
        )


    # 决战top100杯子
    for db_name in get_ini[0]:
        write_top_excel_xlsx(
            db_name, db_name, get_ini[8], get_ini[7], get_ini[2], get_ini[3], 1102
        )

    # 决战
    for db_name in get_ini[0]:
        write_excel_xlsx(
            db_name, db_name, get_ini[1], get_ini[5], get_ini[2], get_ini[3], 1102
        )

    write_log("zhixing wan cheng")

